import tkinter as tk
from tkinter import ttk, font, filedialog
import shutil
import openpyxl
import os
import datetime
from common.version import __version__ as app_version
from common.version import __build_date__ as app_date
from common.constants import GAME_NUMBER_TEMPLATE_PATH
from utils.file_operations import download_template_file

class SortOptionsWindow(tk.Toplevel):
    def __init__(self, master, data):
        super().__init__(master)
        self.title("정렬 순서 변경")
        self.geometry("550x600")
        self.transient(master)
        self.grab_set()

        self.original_data = data
        self.result = None  # 최종 결과를 담을 변수

        # 상단 컨트롤 프레임
        top_controls_frame = tk.Frame(self)
        top_controls_frame.pack(fill=tk.X, padx=10, pady=5)

        # 버튼들을 담을 왼쪽 프레임
        left_button_frame = tk.Frame(top_controls_frame)
        left_button_frame.pack(side=tk.LEFT)

        move_up_button = tk.Button(left_button_frame, text="▲ 위로 이동", command=self._move_up)
        move_up_button.pack(side=tk.LEFT, padx=(0, 5))

        move_down_button = tk.Button(left_button_frame, text="▼ 아래로 이동", command=self._move_down)
        move_down_button.pack(side=tk.LEFT)

        # 체크박스를 담을 오른쪽 프레임
        right_check_frame = tk.Frame(top_controls_frame)
        right_check_frame.pack(side=tk.RIGHT)

        self.sort_by_schedule_var = tk.BooleanVar(value=False)
        sort_by_schedule_check = tk.Checkbutton(right_check_frame, text="일정안 순 정렬", variable=self.sort_by_schedule_var)
        sort_by_schedule_check.pack()

        # Treeview (데이터 표시)
        tree_frame = tk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10)

        self.tree = ttk.Treeview(tree_frame, columns=("순서", "종목", "부", "체급"), show="headings", selectmode='extended')
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)

        self.tree.heading("순서", text="순서")
        self.tree.column("순서", width=50, anchor='center')
        self.tree.heading("종목", text="종목")
        self.tree.column("종목", width=150)
        self.tree.heading("부", text="부")
        self.tree.column("부", width=100)
        self.tree.heading("체급", text="체급")
        self.tree.column("체급", width=100)

        self.item_map = {} # Treeview item ID와 원본 데이터를 매핑
        self._populate_tree()

        # 하단 버튼 프레임
        bottom_button_frame = tk.Frame(self)
        bottom_button_frame.pack(pady=10)

        save_button = tk.Button(bottom_button_frame, text="저장", command=self._save_and_close)
        save_button.pack(side=tk.LEFT, padx=10)

        cancel_button = tk.Button(bottom_button_frame, text="취소", command=self._cancel)
        cancel_button.pack(side=tk.LEFT, padx=10)

    def _populate_tree(self):
        for i, data_row in enumerate(self.original_data):
            values = (i + 1, data_row['종목'], data_row['부'], data_row['체급'])
            item_id = self.tree.insert("", "end", values=values)
            self.item_map[item_id] = data_row

    def _move_up(self):
        selected_items = self.tree.selection()
        if not selected_items:
            return
        sorted_selection = sorted(selected_items, key=lambda x: self.tree.index(x))
        for item_id in sorted_selection:
            current_index = self.tree.index(item_id)
            if current_index > 0:
                self.tree.move(item_id, "", current_index - 1)

    def _move_down(self):
        selected_items = self.tree.selection()
        if not selected_items:
            return
        sorted_selection = sorted(selected_items, key=lambda x: self.tree.index(x), reverse=True)
        total_items = len(self.tree.get_children())
        for item_id in sorted_selection:
            current_index = self.tree.index(item_id)
            if current_index < total_items - 1:
                self.tree.move(item_id, "", current_index + 1)

    def _save_and_close(self):
        new_order_list = []
        for item_id in self.tree.get_children():
            new_order_list.append(self.item_map[item_id])
        
        self.result = {
            "order": new_order_list,
            "sort_by_schedule": self.sort_by_schedule_var.get()
        }
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


class GameNumberCalculator(tk.Toplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.version = app_version
        self.title(f"경기번호 계산기 v{app_version} (빌드: {app_date})")
        self.geometry("1200x700")
        self.last_imported_filename = ""
        self.sort_by_schedule = False # 일정안 순 정렬 옵션

        # 설명 레이블
        description_font = font.Font(family="Helvetica", size=12)
        description_label = tk.Label(self, text="컷오프 계산시 종목에 '자유품새'를 입력하세요.", font=description_font, pady=10)
        description_label.pack()

        # 메인 프레임 (좌우 분할)
        main_container = tk.Frame(self)
        main_container.pack(fill=tk.BOTH, expand=True)

        # 왼쪽: 입력 컨테이너
        input_container = tk.Frame(main_container)
        input_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 버튼 및 헤더 프레임 (입력 컨테이너 내 상단)
        top_controls_frame = tk.Frame(input_container)
        top_controls_frame.pack(fill=tk.X, pady=5)

        # '+' 버튼 (왼쪽 정렬)
        add_row_button = tk.Button(top_controls_frame, text="+ 행추가", command=self.add_row)
        add_row_button.pack(side=tk.LEFT, padx=5)

        # 초기화 버튼 (빨간색 바탕, 흰 글씨)
        reset_button = tk.Button(top_controls_frame, text="초기화", bg="red", fg="white", command=self.reset_all)
        reset_button.pack(side=tk.LEFT, padx=5)

        # 오른쪽 정렬 버튼들을 담을 프레임
        right_buttons_frame = tk.Frame(top_controls_frame)
        right_buttons_frame.pack(side=tk.RIGHT)

        # 계산하기 버튼 (3번째, 붉은색)
        calculate_button = tk.Button(right_buttons_frame, text="계산하기", bg="red", fg="white", command=self.calculate_matches)
        calculate_button.pack(side=tk.RIGHT)

        # 엑셀로 가져오기 버튼 (2번째, 녹색)
        import_button = tk.Button(right_buttons_frame, text="엑셀로 가져오기", bg="#4CAF50", fg="white", command=self.import_from_excel)
        import_button.pack(side=tk.RIGHT, padx=5)

        # 엑셀 양식 다운로드 버튼 (1번째, 기본색)
        download_button = tk.Button(right_buttons_frame, text="엑셀 양식 다운로드", command=self.download_template)
        download_button.pack(side=tk.RIGHT, padx=5)

        # 헤더 프레임 (입력 컨테이너 내 상단, 버튼 아래)
        header_frame = tk.Frame(input_container)
        header_frame.pack(fill=tk.X, pady=5)

        header_labels = ["번호", "종목", "부", "체급", "참가인원"]
        for i, label_text in enumerate(header_labels):
            width = 5 if label_text == "번호" else 15
            label = tk.Label(header_frame, text=label_text, width=width)
            label.pack(side=tk.LEFT, padx=5)

        # '+' 버튼을 위한 빈 공간 (헤더와 정렬)
        tk.Label(header_frame, width=5).pack(side=tk.LEFT, padx=5)

        # 스크롤 가능한 입력 행 프레임
        self.canvas = tk.Canvas(input_container)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar = ttk.Scrollbar(input_container, orient="vertical", command=self.canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.bind('<Configure>', lambda e: self.canvas.configure(scrollregion = self.canvas.bbox("all")))

        self.main_frame = tk.Frame(self.canvas)
        self.canvas.bind('<Configure>', lambda e: self.canvas.itemconfig(self.canvas.create_window((0, 0), window=self.main_frame, anchor="nw", width=e.width), width=e.width))

        self.main_frame.bind('<Configure>', lambda e: self.canvas.configure(scrollregion = self.canvas.bbox("all")))

        # 오른쪽: 결과 프레임
        result_frame = tk.Frame(main_container, width=700)
        result_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)
        result_frame.pack_propagate(False) #결과 프레임 크기 고정

        result_header_frame = tk.Frame(result_frame)
        result_header_frame.pack(fill=tk.X)

        result_label = tk.Label(result_header_frame, text="계산 결과")
        result_label.pack(side=tk.LEFT, pady=5, padx=5)

        result_move_buttons_frame = tk.Frame(result_header_frame)
        result_move_buttons_frame.pack(side=tk.LEFT, padx=10)

        move_up_button = tk.Button(result_move_buttons_frame, text="▲", command=self._move_result_up)
        move_up_button.pack(side=tk.LEFT, padx=2)

        move_down_button = tk.Button(result_move_buttons_frame, text="▼", command=self._move_result_down)
        move_down_button.pack(side=tk.LEFT, padx=2)

        sort_options_button = tk.Button(result_header_frame, text="⚙️ 정렬옵션", command=self._open_sort_options)
        sort_options_button.pack(side=tk.RIGHT, pady=5, padx=5)

        # 결과 표시 Treeview
        tree_frame = tk.Frame(result_frame) # Treeview와 스크롤바를 담을 프레임
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.result_tree = ttk.Treeview(tree_frame, columns=("번호", "종목", "부", "체급", "강수", "경기번호", "경기수"), show='tree headings', selectmode='extended')
        self.result_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tree_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.result_tree.yview)
        tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_tree.configure(yscrollcommand=tree_scrollbar.set)

        # 결과 다운로드 버튼
        # 결과 다운로드 버튼
        download_results_button = tk.Button(result_frame, text="결과 다운로드", bg="red", fg="white", font=("Helvetica", 12, "bold"), command=self.export_results_to_excel)
        download_results_button.pack(fill=tk.X, pady=5)

        # 각 열 설정
        self.result_tree.heading("번호", text="번호")
        self.result_tree.column("번호", width=40, anchor='center')
        self.result_tree.column("#0", width=0, stretch=tk.NO) # 트리 열 숨기기
        self.result_tree.heading("종목", text="종목")
        self.result_tree.column("종목", width=120)
        self.result_tree.heading("부", text="부")
        self.result_tree.column("부", width=80)
        self.result_tree.heading("체급", text="체급")
        self.result_tree.column("체급", width=80)
        self.result_tree.heading("강수", text="강수")
        self.result_tree.column("강수", width=70, anchor='center')
        self.result_tree.heading("경기번호", text="경기번호")
        self.result_tree.column("경기번호", width=80, anchor='center')
        self.result_tree.heading("경기수", text="경기수")
        self.result_tree.column("경기수", width=40, anchor='center')

        # 정렬 상태 초기화
        self.sort_state = {col: 0 for col in self.result_tree["columns"]}

        # 헤더 클릭 이벤트 바인딩
        for col in self.result_tree["columns"]:
            self.result_tree.heading(col, command=lambda _col=col: self._sort_column(_col))

        self.rows = []
        for _ in range(10):
            self.add_row()

        # Treeview에 복사 기능 바인딩
        self.result_tree.bind("<Control-c>", self._copy_selected_rows)
        self.result_tree.bind("<Command-c>", self._copy_selected_rows) # For macOS

        # 마우스 휠 스크롤 바인딩
        self.canvas.bind("<MouseWheel>", self._on_mousewheel) # Windows/Linux
        self.canvas.bind("<Button-4>", self._on_mousewheel) # macOS scroll up
        self.result_tree.bind("<Button-5>", self._on_mousewheel) # macOS scroll down

        # Footer
        footer_font = font.Font(family="Helvetica", size=9)
        footer_label = tk.Label(self, text="Copyright (c) FEELJAE-WON. All rights reserved.", font=footer_font, fg="gray")
        footer_label.pack(side=tk.BOTTOM, pady=5)
        self.main_frame.bind("<MouseWheel>", self._on_mousewheel) # Windows/Linux
        self.main_frame.bind("<Button-4>", self._on_mousewheel) # macOS scroll up
        self.main_frame.bind("<Button-5>", self._on_mousewheel) # macOS scroll down

    def _open_sort_options(self):
        # 1. 현재 입력된 데이터 가져오기
        current_data = []
        for row_data in self.rows:
            entries = row_data["entries"]
            event = entries["종목"].get().strip()
            division = entries["부"].get().strip()
            weight_class = entries["체급"].get().strip()
            participants = entries["참가인원"].get().strip()

            # 참가인원이 있는 행만 정렬 대상으로 포함
            if participants:
                current_data.append({
                    "종목": event,
                    "부": division,
                    "체급": weight_class,
                    "참가인원": participants,
                    "original_row_data": row_data # 원래 위젯 참조를 저장
                })

        if not current_data:
            tk.messagebox.showinfo("정보", "정렬할 데이터가 없습니다.")
            return

        # 2. 정렬 옵션 창 열기
        sort_window = SortOptionsWindow(self, current_data)
        self.wait_window(sort_window) # 창이 닫힐 때까지 대기

        # 3. 변경된 순서와 옵션 적용
        if sort_window.result:
            self.sort_by_schedule = sort_window.result["sort_by_schedule"]
            if sort_window.result["order"]:
                self._apply_new_order(sort_window.result["order"])

    def _apply_new_order(self, new_order):
        # 새로운 순서에 맞게 기존 입력 필드의 값을 업데이트합니다.
        # new_order 리스트의 순서대로 self.rows의 해당 위젯에 값을 재설정합니다.

        # 비어있지 않은 행들만 가져옵니다.
        active_rows = [row for row in self.rows if row["entries"]["참가인원"].get().strip()]

        # new_order의 순서대로 active_rows의 위젯에 값을 채워넣습니다.
        for i, sorted_data in enumerate(new_order):
            if i < len(active_rows):
                target_row_widgets = active_rows[i]["entries"]
                target_row_widgets["종목"].delete(0, tk.END)
                target_row_widgets["종목"].insert(0, sorted_data["종목"])
                target_row_widgets["부"].delete(0, tk.END)
                target_row_widgets["부"].insert(0, sorted_data["부"])
                target_row_widgets["체급"].delete(0, tk.END)
                target_row_widgets["체급"].insert(0, sorted_data["체급"])
                target_row_widgets["참가인원"].delete(0, tk.END)
                target_row_widgets["참가인원"].insert(0, sorted_data["참가인원"])

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def _copy_selected_rows(self, event=None):
        selected_items = self.result_tree.selection()
        if not selected_items:
            return

        clipboard_content = []
        for item_id in selected_items:
            values = self.result_tree.item(item_id, 'values')
            clipboard_content.append('\t'.join(map(str, values)))
        
        self.clipboard_clear()
        self.clipboard_append('\n'.join(clipboard_content))

    def reset_all(self):
        # 결과 Treeview 초기화
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)

        # 입력 필드 초기화
        for row_data in self.rows:
            row_data["frame"].destroy()
        self.rows.clear()

        # 초기 10개 행 추가
        for _ in range(10):
            self.add_row()

    def remove_row(self, row_frame_to_remove):
        for i, row_data in enumerate(self.rows):
            if row_data["frame"] == row_frame_to_remove:
                row_data["frame"].destroy()
                del self.rows[i]
                self.resequence_row_numbers()
                break

    def resequence_row_numbers(self):
        for i, row_data in enumerate(self.rows):
            row_data["number_label"].config(text=str(i + 1))

    def add_row(self):
        row_frame = tk.Frame(self.main_frame)
        row_frame.pack(fill=tk.X, expand=True, pady=5)

        row_number = len(self.rows) + 1
        number_label = tk.Label(row_frame, text=str(row_number), width=5)
        number_label.pack(side=tk.LEFT, padx=5)

        labels = ["종목", "부", "체급", "참가인원"]
        entries = {}

        for i, label_text in enumerate(labels):
            entry = tk.Entry(row_frame, width=15)
            entry.pack(side=tk.LEFT, padx=5)
            entries[label_text] = entry
        
        # '-' 버튼 추가
        remove_button = tk.Button(row_frame, text="-", command=lambda r_frame=row_frame: self.remove_row(r_frame))
        remove_button.pack(side=tk.LEFT, padx=5)

        self.rows.append({"frame": row_frame, "entries": entries, "number_label": number_label, "remove_button": remove_button})

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls *"), ("All files", "*.*")])
        if not file_path:
            return

        self.last_imported_filename = os.path.splitext(os.path.basename(file_path))[0]

        # Clear existing rows
        for row_data in self.rows:
            row_data["frame"].destroy()
        self.rows.clear()

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0: # Skip header row
                continue
            self.add_row_with_data(row)

    def download_template(self):
        download_template_file(GAME_NUMBER_TEMPLATE_PATH, "경기번호_계산기_양식.xlsx", [("Excel files", "*.xlsx")])

    def calculate_matches(self):
        # 기존 결과 삭제
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)

        # 정렬 상태 초기화 및 헤더 화살표 제거
        for col in self.result_tree["columns"]:
            self.sort_state[col] = 0
            current_text = self.result_tree.heading(col, "text")
            self.result_tree.heading(col, text=current_text.replace(" ▲", "").replace(" ▼", ""))

        game_number_counter = 1
        row_index = 1 # 결과 테이블의 행 번호
        prev_event, prev_division, prev_weight_class = None, None, None

        for i, row_data in enumerate(self.rows):
            entries = row_data["entries"]
            try:
                participants_str = entries["참가인원"].get()
                if not participants_str.strip():
                    continue

                participants = int(participants_str)
                if participants < 2:
                    continue

                event = entries["종목"].get() or ""
                division = entries["부"].get() or ""
                weight_class = entries["체급"].get() or ""

                if event == "자유품새":
                    if participants <= 11:
                        # Case 1: Participants <= 11
                        self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, "결선", f"1~{participants}", participants))
                        row_index += 1
                    elif 12 <= participants <= 21:
                        # Case 2: 12 <= Participants <= 21
                        # Divide into 2 groups, with the first group being larger if uneven
                        group1_size = (participants + 1) // 2
                        group2_size = participants - group1_size

                        # 본선-1조
                        self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, "본선-1조", f"1~{group1_size}", group1_size))
                        row_index += 1
                        # 본선-2조
                        self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, "본선-2조", f"1~{group2_size}", group2_size))
                        row_index += 1

                        # 결선
                        self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, "결선", "1~8", 8))
                        row_index += 1
                    elif participants >= 22:
                        # Case 3: Participants >= 22
                        # Preliminary (예선)
                        num_prelim_groups = 2 # Start with 2 groups
                        while participants / num_prelim_groups > 11.5 and num_prelim_groups % 2 == 0:
                            num_prelim_groups += 2

                        # Ensure num_prelim_groups is at least 2 and even
                        if num_prelim_groups < 2:
                            num_prelim_groups = 2
                        if num_prelim_groups % 2 != 0:
                            num_prelim_groups += 1

                        base_prelim_group_size = participants // num_prelim_groups
                        remainder_prelim = participants % num_prelim_groups

                        prelim_group_sizes = []
                        for g in range(num_prelim_groups):
                            size = base_prelim_group_size
                            if g < remainder_prelim:
                                size += 1
                            prelim_group_sizes.append(size)

                        # 예선 결과 출력
                        for g_idx, size in enumerate(prelim_group_sizes):
                            self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, f"예선-{g_idx+1}조", f"1~{size}", size))
                            row_index += 1

                        # 본선 진출 인원 계산 (1조 인원의 절반(반올림)이 해당 체급에 모든 조에 그 인원이 진출)
                        first_group_size = prelim_group_sizes[0] if prelim_group_sizes else 0
                        advancement_per_group = (first_group_size + 1) // 2 # Round up for advancement
                        total_main_round_participants = advancement_per_group * num_prelim_groups

                        # 본선 (Main Round)
                        if total_main_round_participants > 0:
                            num_main_groups = 2 # Start with 2 groups
                            while total_main_round_participants / num_main_groups > 11.5 and num_main_groups % 2 == 0:
                                num_main_groups += 2

                            # Ensure num_main_groups is at least 2 and even
                            if num_main_groups < 2:
                                num_main_groups = 2
                            if num_main_groups % 2 != 0:
                                num_main_groups += 1

                            base_main_group_size = total_main_round_participants // num_main_groups
                            remainder_main = total_main_round_participants % num_main_groups

                            main_group_sizes = []
                            for g in range(num_main_groups):
                                size = base_main_group_size
                                if g < remainder_main:
                                    size += 1
                                main_group_sizes.append(size)

                            # 본선 결과 출력
                            for g_idx, size in enumerate(main_group_sizes):
                                self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, f"본선-{g_idx+1}조", f"1~{size}", size))
                                row_index += 1

                            # 결선 (Final Round) - based on number of main round groups
                            # 예선과 본선을 거쳤을 경우 결선은 무조건 1~8
                            self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, "결선", "1~8", 8))
                            row_index += 1
                    
                    prev_event, prev_division, prev_weight_class = None, None, None

                else:
                    # Existing logic for other events
                    current_category = (event, division, weight_class)
                    previous_category = (prev_event, prev_division, prev_weight_class)

                    if current_category != previous_category:
                        game_number_counter = 1

                    game_number_counter, row_index = self._calculate_standard_matches(participants, event, division, weight_class, game_number_counter, row_index)
                    
                    prev_event, prev_division, prev_weight_class = event, division, weight_class

            except ValueError:
                continue

        # "일정안 순 정렬" 옵션이 켜져 있으면 특별 정렬 수행
        if self.sort_by_schedule:
            self._sort_results_by_schedule()

    def _calculate_standard_matches(self, participants, event, division, weight_class, game_number_counter, row_index):
        total_slots = 1
        while total_slots < participants:
            total_slots *= 2

        byes = total_slots - participants
        first_round_matches = participants - byes

        # 예선전 (첫 라운드)
        if first_round_matches > 0:
            round_name = f"{total_slots}"
            num_matches = first_round_matches // 2
            start_game = game_number_counter
            end_game = game_number_counter + num_matches - 1
            if num_matches > 0:
                if start_game == end_game:
                    game_numbers_display = f"{start_game}"
                else:
                    game_numbers_display = f"{start_game}~{end_game}"
            else:
                game_numbers_display = "-"
            self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, round_name, game_numbers_display, num_matches))
            game_number_counter = end_game + 1
            row_index += 1

        # 본선 (다음 라운드부터 결승까지)
        current_participants = (first_round_matches // 2) + byes
        while current_participants > 1:
            round_matches = current_participants // 2
            round_name = f"{current_participants}" if current_participants > 2 else "2"
            if current_participants == 4:
                round_name = "4"
            elif current_participants == 8:
                round_name = "8"

            start_game = game_number_counter
            end_game = game_number_counter + round_matches - 1
            if round_matches > 0:
                if start_game == end_game:
                    game_numbers_display = f"{start_game}"
                else:
                    game_numbers_display = f"{start_game}~{end_game}"
            else:
                game_numbers_display = "-"
            self.result_tree.insert("", "end", values=(row_index, event, division, weight_class, round_name, game_numbers_display, round_matches))
            game_number_counter = end_game + 1
            current_participants //= 2
            row_index += 1
        return game_number_counter, row_index

    def _sort_column(self, col):
        # 현재 열의 정렬 상태 업데이트
        current_state = self.sort_state[col]
        
        # 모든 헤더에서 화살표 제거
        for c in self.result_tree["columns"]:
            current_text = self.result_tree.heading(c, "text")
            self.result_tree.heading(c, text=current_text.replace(" ▲", "").replace(" ▼", ""))

        if current_state == 0: # 정렬 안됨 -> 내림차순
            new_state = 1
            reverse = True
            arrow = " ▼"
        elif current_state == 1: # 내림차순 -> 오름차순
            new_state = 2
            reverse = False
            arrow = " ▲"
        else: # 오름차순 -> 정렬 취소
            new_state = 0
            reverse = False # 정렬 취소 시에는 순서 무의미
            arrow = ""

        self.sort_state = {c: 0 for c in self.result_tree["columns"]} # 모든 열 정렬 상태 초기화
        self.sort_state[col] = new_state

        # 현재 열 헤더에 화살표 추가
        current_text = self.result_tree.heading(col, "text")
        self.result_tree.heading(col, text=current_text.split(" ")[0] + arrow)

        # 데이터 가져오기
        data = []
        for item_id in self.result_tree.get_children():
            data.append((self.result_tree.item(item_id, 'values'), item_id))

        # 정렬
        if new_state != 0:
            col_index = self.result_tree["columns"].index(col)
            if col == "강수":
                data.sort(key=lambda x: self._get_round_value(x[0][col_index]), reverse=reverse)
            else:
                data.sort(key=lambda x: x[0][col_index], reverse=reverse)
            #data.sort(key=lambda x: x[0][col_index], reverse=reverse)

        # Treeview 업데이트
        for item_id in self.result_tree.get_children():
            self.result_tree.delete(item_id)

        for idx, (values, item_id) in enumerate(data):
            # '번호' 열을 현재 순서에 맞게 업데이트
            updated_values = list(values)
            updated_values[0] = idx + 1
            self.result_tree.insert('', 'end', values=updated_values)

    def _get_round_value(self, round_str):
        # Custom sort for "자유품새" rounds
        if '결선' in round_str:
            # Highest priority for ascending sort
            return (0, 0)
        elif '본선' in round_str:
            try:
                group_num = int(round_str.split('-')[1].replace('조', ''))
                # Second priority
                return (1, group_num)
            except (IndexError, ValueError):
                return (1, 0) # Fallback for "본선" without a group number
        elif '예선' in round_str:
            try:
                group_num = int(round_str.split('-')[1].replace('조', ''))
                # Third priority
                return (2, group_num)
            except (IndexError, ValueError):
                return (2, 0) # Fallback for "예선" without a group number

        # Sort logic for standard tournament rounds (e.g., "8", "4")
        try:
            # For ascending sort, smaller numbers come first.
            # For descending, larger numbers come first.
            # This is the natural integer order.
            return (3, int(round_str))
        except ValueError:
            # Fallback for any other string that doesn't fit the patterns above
            return (4, round_str)

    def _sort_results_by_schedule(self):
        all_items = self.result_tree.get_children("")
        if not all_items:
            return

        # 데이터와 정렬 키를 함께 저장
        data_to_sort = []
        for item in all_items:
            values = self.result_tree.item(item, 'values')
            kangsoo = self._get_kangsoo_value(values[4]) # 강수
            game_number_str = str(values[5]) # 경기번호
            
            # 해당 강수에서 새로 시작하는 경기인지 (경기번호가 1로 시작)
            is_new_start = game_number_str.startswith('1') or game_number_str.startswith('1~')

            # 정렬 기준: 1. 강수(내림차순), 2. 새로운 시작 여부(True가 먼저), 3. 원래 순서(번호 오름차순)
            sort_key = (-kangsoo, not is_new_start, int(values[0]))
            data_to_sort.append((sort_key, values))

        # 정렬 실행
        data_to_sort.sort(key=lambda x: x[0])

        # Treeview 비우기
        for item in all_items:
            self.result_tree.delete(item)

        # 정렬된 데이터로 Treeview 다시 채우기
        for i, (key, values) in enumerate(data_to_sort):
            new_values = list(values)
            new_values[0] = i + 1 # 번호 재지정
            self.result_tree.insert("", "end", values=new_values)

    def _get_kangsoo_value(self, kangsoo_str):
        try:
            # "128강", "64" 등 숫자 부분만 추출하여 정수로 변환
            return int("".join(filter(str.isdigit, str(kangsoo_str))))
        except:
            # 변환 실패 시 낮은 우선순위 부여
            return 0

    def export_results_to_excel(self):
        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"경기번호계산_{current_time}.xlsx"

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                   initialfile=default_filename,
                                                   filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "경기 결과"

        # 헤더 추가
        headers = [self.result_tree.heading(col, "text").replace(" ▲", "").replace(" ▼", "") for col in self.result_tree["columns"]]
        sheet.append(headers)

        # 데이터 추가
        for item_id in self.result_tree.get_children():
            values = self.result_tree.item(item_id, 'values')
            sheet.append(values)

        try:
            workbook.save(file_path)
            tk.messagebox.showinfo("성공", f"결과가 {file_path}에 저장되었습니다.")
        except Exception as e:
            tk.messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {e}")

    def _move_result_up(self):
        selected_items = self.result_tree.selection()
        if not selected_items:
            return

        sorted_selection = sorted(selected_items, key=lambda x: self.result_tree.index(x))

        for item_id in sorted_selection:
            current_index = self.result_tree.index(item_id)
            if current_index > 0:
                self.result_tree.move(item_id, "", current_index - 1)
        
        self._resequence_result_numbers()

    def _move_result_down(self):
        selected_items = self.result_tree.selection()
        if not selected_items:
            return

        sorted_selection = sorted(selected_items, key=lambda x: self.result_tree.index(x), reverse=True)

        total_items = len(self.result_tree.get_children())
        for item_id in sorted_selection:
            current_index = self.result_tree.index(item_id)
            if current_index < total_items - 1:
                self.result_tree.move(item_id, "", current_index + 1)

        self._resequence_result_numbers()

    def _resequence_result_numbers(self):
        all_items = self.result_tree.get_children()
        for i, item_id in enumerate(all_items):
            current_values = self.result_tree.item(item_id, 'values')
            new_values = list(current_values)
            new_values[0] = i + 1
            self.result_tree.item(item_id, values=new_values)

    def add_row_with_data(self, data):
        row_frame = tk.Frame(self.main_frame)
        row_frame.pack(fill=tk.X, expand=True, pady=5)

        row_number = len(self.rows) + 1
        number_label = tk.Label(row_frame, text=str(row_number), width=5)
        number_label.pack(side=tk.LEFT, padx=5)

        labels = ["종목", "부", "체급", "참가인원"]
        entries = {}

        for i, label_text in enumerate(labels):
            entry = tk.Entry(row_frame, width=15)
            entry.pack(side=tk.LEFT, padx=5)
            if i < len(data):
                # Convert None to empty string
                value = "" if data[i] is None else data[i]
                entry.insert(0, value)
            entries[label_text] = entry

        # '-' 버튼 추가
        remove_button = tk.Button(row_frame, text="-", command=lambda r_frame=row_frame: self.remove_row(r_frame))
        remove_button.pack(side=tk.LEFT, padx=5)

        self.rows.append({"frame": row_frame, "entries": entries, "number_label": number_label, "remove_button": remove_button})

if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    app = GameNumberCalculator(master=root)
    app.mainloop()
